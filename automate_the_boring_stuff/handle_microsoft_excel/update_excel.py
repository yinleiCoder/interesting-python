import openpyxl
from openpyxl.styles import Font
"""
更新电子表格:
    
"""
wb = openpyxl.load_workbook('2021换证考试成绩表1.xlsx')
sheet = wb.active
SCORE_UPDATES = {'赵飞': 100}
sheet.freeze_panes = 'A3'
for rowNum in range(2, sheet.max_row):
    stuName = sheet.cell(row=rowNum, column=2).value
    if stuName in SCORE_UPDATES:
        sheet.cell(row=rowNum, column=8).value = SCORE_UPDATES[stuName]
        boldFont = Font(24, bold=True)
        # list(sheet.rows)[rowNum]
        sheet.cell(row=rowNum, column=2).font = boldFont
        sheet.row_dimensions[rowNum].height = 100
        sheet.merge_cells('A3:H3')
        sheet.unmerge_cells('A3:H3')

sheet.cell(row=sheet.max_row+1, column=1).value = f'=SUM(H3:H{sheet.max_row})'

refObj = openpyxl.chart.Reference(sheet, min_col=8, min_row=1, max_col=8, max_row=sheet.max_row)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'Score Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'J3')


wb.save('2021换证考试改赵飞室友成绩.xlsx')